import pymongo
import oop
from colorama import init,Fore,Style
init(autoreset=True)
import string
import random
import hashlib
from openpyxl import Workbook
from datetime import datetime
import os
import time


letters = string.ascii_letters
nums = string.digits
id_stuff = nums + string.ascii_lowercase
sings = string.punctuation
stuffs = letters + nums + sings

def get_connection() :
    con = pymongo.MongoClient("mongodb://localhost:27017")
    db = con["fainace_manger"]
    return db

def creat_p():
    password = "".join(random.sample(stuffs, 10))
    return password

def create_acc():
    db = get_connection()
    col = db["users"]
    us = oop.person()
    us.name = input("Enter your name : ")
    us.family = input("Enter your family : ")
    while True:
        
        us.mobile = input("Enter your mobile number: ")
        m = us.mobile
        valid_local = m.startswith("09") and len(m) == 11 and m.isdigit()
        valid_inter = m.startswith("+98") and len(m) == 13 and m[1:].isdigit()
        
        if valid_local or valid_inter:
            mob_lst = list(col.find({"mobile" : us.mobile}))
            #print(mob_lst)
            if mob_lst == [] :
                break
            
            if mob_lst[0].get("mobile") == us.mobile:
                print(Fore.RED+"This mobile number is used.")
        else:
            print(Fore.RED+"Invalid mobile number. Try again.")
    
    us.username = input("Enter your username : ")
    user_lst = list(col.find({"username" : us.username}))
    while True :
        if user_lst == [] :
            break
        
        if user_lst[0].get("username") == us.username : 
            print("This username is already taken try new one")
            us.username = input("Enter your username : ")
    us.password = creat_p().strip()
    print(f"Username = {us.username},Password = {us.password}")
    
    us.password = hashlib.sha256(us.password.encode()).hexdigest()

    
    data = {
        "ID":us.id,"Name" : us.name,"Family" : us.family,"mobile" : us.mobile,
        "username" : us.username,"password" :us.password
    }
    col.insert_one(data)
    print("Account created")

def login():
    db = get_connection()
    col = db["users"]
    i = 0
    while(i<=5) :
        i = i + 1
        print(i)
        username = input("Enter your username : ")
        password = input("Enter your password : ")
        password = hashlib.sha256(password.encode()).hexdigest()
        user = col.find_one({"$and":[{"username": username},{"password": password}]})
       
        if user:
            id = user.get("ID")
            print(id)
            name = f"{user.get('Name')} {user.get('Family')}"
            print(f"{Fore.GREEN}Welcome {name}{Style.RESET_ALL}")
            while True:
                print("ac --> add card")
                print("c  --> charge card")
                print("sc --> see card")
                print("rc --> remove card")
                print("at --> add transaction")
                print("st --> see tarakonesh")
                print("bt --> backup tarakonesh")
                print("e  --> exit")
                ch = input("Enter your choice : ").lower()
                if ch == "ac":
                    add_card(id, name)
                elif ch == "sc":
                    see_card(id)
                elif ch == "c":
                    Ccharge(id)
                elif ch == "rc" :
                    remove_card(id)
                elif ch == "at":
                    add_tar(id)
                elif ch == "st" :
                    see_tar(id)
                elif ch == "bt" :
                    tbackup(id)
                elif ch == "e":
                    print("Goodbye")
                    return
                else:
                    print(Fore.YELLOW + "Invalid option. Try again.")
            
        else:
            print(Fore.RED + "Wrong username or password" + Style.RESET_ALL)
            i += 1
            if i >= 5:
                    print(Fore.RED + "Too many failed attempts. Please wait 60 seconds..." + Style.RESET_ALL)
                    time.sleep(60)
                    i = 0
                   
def add_card(id,name) :
    db = get_connection()
    col = db["card"]
    ca = oop.card()
    while True :
        ca.Cnumber = input("Enter your card number : ")
        if list(col.find({"card number" : ca.Cnumber})) != []:
            print("This card already add")
            break
        else :
            while True :
                if ca.Cnumber[:6] in oop.banks and len(ca.Cnumber) == 16:
                    break
                else :
                    print(Fore.RED+"try again")
                    ca.Cnumber = input("Enter your card number : ")
            exmonth = input("Enter card expier month(03 or 10) : ")
            exyear = input("Enter card expier year : ")
            ca.expdate = f"{exmonth}/{exyear}"
            ca.cvv2 = input("Enter card cvv2 : ")
            ca.bankName =  oop.banks_name.get(ca.Cnumber[:6])
            while True :
                if (len(ca.cvv2) == 3 or len(ca.cvv2) == 4) and ca.cvv2.isdigit() :
                    break
                else :
                    print(Fore.RED+"try again")
                    ca.cvv2 = input("Enter card cvv2 : ")
            
            data = {"card ID" :ca.cardId,"ownerID":id,"card number" : ca.Cnumber,
                    "expier date" : ca.expdate,"CVV2":ca.cvv2,"Ownername" : name,
                    "Bank" : ca.bankName,"value" :ca.value}
            col.insert_one(data)
            print("Card add")
            break
    
def see_card(id) :
    db = get_connection()
    col = db["card"]
    data = list(col.find({"ownerID" : id}))
    if data != [] :
        for i in range(len(data)) :
            Cinfo = list(data[i].values())
            print(f"{i+1}- Cardnumber : {Cinfo[3]}  Bank : {Cinfo[-2]}  Value : {Cinfo[-1]}  ID :{Cinfo[1]}")
            print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
    else :
        print(Fore.RED+"You dont have any card" + Style.RESET_ALL)
   
def remove_card(id) :
    db = get_connection()
    col = db["card"]
    col2 = db["tarakonesh"]
    data = list(col.find({"ownerID" : id}))
    if data != [] :
        idLst = []
        for i in range(len(data)) :
            Cinfo = list(data[i].values())
            print(f"{i+1}- Cardnumber : {Cinfo[3]}  Bank : {Cinfo[-2]}  Value : {Cinfo[-1]}  ID :{Cinfo[1]}")
            idLst.append(Cinfo[1])
            print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
        Cid = int(input("Enter the card ID : "))
        while True :
            if Cid in idLst :
                att = input("Are you sure for delete this card(Y/N)").lower()
                if att == "y" :
                    col.delete_one({"card ID" : Cid})
                    print("removed from Cards")
                    col2.delete_many({"cardID" : Cid})
                    print("removed from tarakonesh")
                    break
                elif att == "n" :
                    print("OK")
                    break
                
            else :
                print(Fore.RED+"Invalid ID")
                Cid = int(input("Enter the card ID : "))

    else :
        print(Fore.RED+"You dont have any card" + Style.RESET_ALL)
       
def Ccharge(id) :
    db = get_connection() 
    col = db["card"]
    data = list(col.find({"ownerID" : id}))
    if data != [] :
        idLst = []
        for i in range(len(data)) :
            Cinfo = list(data[i].values()) 
            print(f"{i+1}- Cardnumber : {Cinfo[3]}  Bank : {Cinfo[-2]}  Value : {Cinfo[-1]}  ID :{Cinfo[1]}") 
            print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
            idLst.append(Cinfo[1])   
        #print(idLst)
        Cid = int(input("Enter the card id : "))
        while True :
            if Cid not in idLst : 
                print("Wrong id")
                Cid = input("Enter the card id : ")
            if Cid in idLst :
                value = int(input("Enter amount of you want charge : "))
                col.find_one_and_update({"card ID" : Cid},{"$inc" :{"value":value}})
                print("Charge Done")
                break
    else :
        print(Fore.RED+"You dont have any card" + Style.RESET_ALL)
           
def add_tar(id) :
    db = get_connection() 
    col = db["card"]
    col2 = db["tarakonesh"]
    data = list(col.find({"ownerID" : id}))
    if data != [] :
        idLst = []
        for i in range(len(data)) :
            Cinfo = list(data[i].values()) 
            print(f"{i+1}- Cardnumber : {Cinfo[3]}  Bank : {Cinfo[-2]}  Value : {Cinfo[-1]}  ID :{Cinfo[1]}") 
            print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
            idLst.append(Cinfo[1])   
        Cid = int(input("Enter the card id : "))
        while True :
            if Cid not in idLst : 
                print(Fore.RED+"Wrong id")
                Cid = int(input("Enter the card id : "))
            if Cid in idLst :
                data = list(col.find({"card ID" : Cid}))
                data = list(data[0].values())
                ta = oop.tarakonesh()
                ta.T_type = input("How was your tarakonesh (increase/decrease) : ")
                
                while  not (ta.T_type == "decrease" or ta.T_type == "increase") :
                    print(Fore.RED+"try agian")
                    ta.T_type = input("How was your tarakonesh (increase/decrease) : ")
                    
                ta.T_value = int(input("Enter the value of tarakonesh : "))
                res = input("Do you want add reaseon for tarakonesh(Y/N)").lower()
                if res == "y" :
                    print(oop.tar_cate)
                    ta.reaseon = input("Enter the reaseon : ")
                if res == "n" :
                    ta.reaseon = "NoReaseon"
                cardId = data[1]
                bankName = data[-2]
                if ta.T_type == "increase":
                    col.find_one_and_update({"card ID": Cid}, {"$inc": {"value": ta.T_value}})
                elif ta.T_type == "decrease":
                    col.find_one_and_update({"card ID": Cid}, {"$inc": {"value": -ta.T_value}})
                
                remaining_money = list(col.find({"card ID" : Cid}))
                remaining_money = list(remaining_money[0].values())
                remaining_money = remaining_money[-1]
                Tdata = {
                    "ID" : ta.T_id,"cardID":cardId,"Bank" : bankName,
                    "type" : ta.T_type,"value":ta.T_value,"date" : ta.T_date,
                    "reaseon" : ta.reaseon,"remaining money" : remaining_money
                }
                col2.insert_one(Tdata)
                print("Tarakonesh add")
                print("Done")
                break
    else :
        print(Fore.RED+"You dont have any card" + Style.RESET_ALL)
                               
def see_tar(id) :
    db = get_connection() 
    col = db["card"]
    col2 = db["tarakonesh"]
    data = list(col.find({"ownerID" : id}))
    if data != [] :
        idLst = []
        for i in range(len(data)) :
            Cinfo = list(data[i].values()) 
            print(f"{i+1}- Cardnumber : {Cinfo[3]}  Bank : {Cinfo[-2]}  Value : {Cinfo[-1]}  ID :{Cinfo[1]}") 
            print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
            idLst.append(Cinfo[1])   
        
        Cid = int(input("Enter the card id : "))
        while True :
            if Cid in idLst :
                ch = input("Do you want filter the search(Y/N) : ").lower()
                if ch == "y" :
                    ch3 = input("Do you want filter by reaseon or type or date : ")
                    if ch3 == "reaseon" :
                        reas = input("Enter the reasoen for search : ")
                        if reas in oop.tar_cate :
                            data = list(col2.find({"$and" : [{"cardID" : Cid},{"reaseon" : reas}]}))
                            if data != []:
                                for i in range(len(data)) :
                                    Tdata = list(data[i].values())
                                    print(f"TarakoneshID : {Tdata[1]}  value : {Tdata[5]}  Tarakonesh type : {Tdata[4]}  Date : {Tdata[-3]}  remaining money : {Tdata[-1]} Reaseon : {Tdata[-2]}")
                                    print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
                                break
                            else :
                                print(Fore.RED+"You dont have tarakonesh with this reaseon" + Style.RESET_ALL)   
                                break     
                        else :
                            print(Fore.RED+"Invalid reasoan" + Style.RESET_ALL)    
                    
                    if ch3 == "type" :
                        Ttype = input("Enter the tarakonesh type(increase/decrease) : ")
                        if Ttype == "increase" or Ttype == "decrease" :
                            data = list(col2.find({"$and" : [{"cardID" : Cid},{"type" : Ttype}]}))
                            if data != []:
                                for i in range(len(data)) :
                                    Tdata = list(data[i].values())
                                    print(f"TarakoneshID : {Tdata[1]}  value : {Tdata[5]}  Tarakonesh type : {Tdata[4]}  Date : {Tdata[-3]}  remaining money : {Tdata[-1]} Reaseon : {Tdata[-2]}")
                                    print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
                                break
                            else :
                                print(Fore.RED+"You dont have tarakonesh with this type" + Style.RESET_ALL)   
                                break     
                            
                        else :
                            print(Fore.RED+"Invalid type") 
                    if ch3 == "date" :
                        print("Enter the start date detail")
                        inputyear = int(input("Enter the year : "))
                        inputmonth = int(input("Enter the month : "))
                        inputday = int(input("Enter the day : "))
                        start_date = datetime(year=inputyear,month=inputmonth,day=inputday)
                        endinput = input("Do you want have a end date or no(Y/N) : ").lower()
                        if endinput == "n" : 
                            data = list(col2.find({"$and" : [{"cardID" : Cid},{"date": {"$gte": start_date, "$lt":datetime.now()}}]}))
                            if data != []:
                                for i in range(len(data)) :
                                    Tdata = list(data[i].values())
                                    print(f"TarakoneshID : {Tdata[1]}  value : {Tdata[5]}  Tarakonesh type : {Tdata[4]}  Date : {Tdata[-3]}  remaining money : {Tdata[-1]} Reaseon : {Tdata[-2]}")
                                    print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
                                break
                            else :
                                print(Fore.RED+"You dont have tarakonesh in this range time" + Style.RESET_ALL)   
                                break   
                        if endinput == "y" :
                            endyear = int(input("Enter the end year : "))
                            endmonth = int(input("Enter the end month : "))
                            endday = int(input("Enter the end day : "))  
                            end_date = datetime(year=endyear,month=endmonth,day=endday)
                            data = list(col2.find({"$and" : [{"cardID" : Cid},{"date": {"$gte": start_date, "$lt":end_date}}]}))
                            if data != []:
                                for i in range(len(data)) :
                                    Tdata = list(data[i].values())
                                    print(f"TarakoneshID : {Tdata[1]}  value : {Tdata[5]}  Tarakonesh type : {Tdata[4]}  Date : {Tdata[-3]}  remaining money : {Tdata[-1]} Reaseon : {Tdata[-2]}")
                                    print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
                                break
                            else :
                                print(Fore.RED+"You dont have tarakonesh in this range time" + Style.RESET_ALL)   
                                break   
                            
                        
                    else :
                        print("Invalid input")
                        break
                    
                if ch == "n" :
                    data = list(col2.find({"cardID" : Cid}).limit(10))
                    for i in range(len(data)) :
                        Tdata = list(data[i].values())
                        print(f"TarakoneshID : {Tdata[1]}  value : {Tdata[5]}  Tarakonesh type : {Tdata[4]}  Date : {Tdata[-3]}  remaining money : {Tdata[-1]}")
                        print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
                    ch2 = input("Do you want see all tarakonesh(Y/N) : ").lower()
                    if ch2 == "y":
                        data = list(col2.find({"cardID" : Cid}))
                        for i in range(len(data)) :
                            Tdata = list(data[i].values())
                            print(f"TarakoneshID : {Tdata[1]}  value : {Tdata[5]}  Tarakonesh type : {Tdata[4]}  Date : {Tdata[-3]}  remaining money : {Tdata[-1]}")
                            print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
                        break
                    if ch2 == "n" :
                        break
            else :
                print(Fore.RED+"Wrong id" + Style.RESET_ALL)
                Cid = int(input("Enter the card id : "))
    else :
        print(Fore.RED+"You dont have any card" + Style.RESET_ALL)        
  
def tbackup(id) :
    db = get_connection()
    col = db["card"]
    col2 =db["tarakonesh"]  
    
    w = Workbook()
    ws = w.active
    ws1 = w.create_sheet("Sheet1", 0)
    tit = ["ID","CardID","Bank","Type","value","Date","reaseon","remaining money"]
    for i in range(len(tit)):
        ws1.cell(row=1, column=i + 1).value = tit[i]
    
    data = list(col.find({"ownerID" : id}))
    if data != [] :
        idLst = []
        for i in range(len(data)) :
            Cinfo = list(data[i].values())
            print(f"{i+1}- Cardnumber : {Cinfo[3]}  Bank : {Cinfo[-2]}  Value : {Cinfo[-1]}  ID :{Cinfo[1]}") 
            print(Fore.CYAN+"*"*20 + Style.RESET_ALL)
            idLst.append(Cinfo[1])
        Cid = int(input("Enter the card ID you want : "))
        while True :
            if Cid in idLst :
                data = list(col2.find({"cardID":Cid}).limit(100))
                Tdata = []
                for j in range(len(data)) :
                    tmp = list(data[j].values())
                    Tdata.append(tmp[1:])
                for i in range(len(Tdata)) :
                    for j in range(len(Tdata[i])) :
                        ws1.cell(row=i + 2, column=j + 1).value = Tdata[i][j]
                na = f"Tarakonesh_{datetime.now().year}_{datetime.now().month}_{random.randint(100, 1000)}.xlsx"
                if os.path.exists("Backup") :
                    w.save(filename=rf"D:\Learn_programing\Python\projects\fainec_maneger\Backup\{na}")
                    print(f"Saved at D:\Learn_programing\Python\projects\fainec_maneger\Backup")
                else :
                    os.mkdir("Backup")
                    w.save(filename=rf"D:\Learn_programing\Python\projects\fainec_maneger\Backup\{na}")
                    print(f"Saved at D:\Learn_programing\Python\projects\fainec_maneger\Backup")
                break
            if Cid not in idLst :
                print(Fore.RED+"Worng Card ID")
                Cid = int(input("Enter the card ID you want : "))  
                break 
    else :
        print(Fore.RED+"You dont have any card")
    

login()
